import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_ebitda_bridge_excel(consolidated_data, filename="EBITDA_Bridge_Report.xlsx"):
    """
    Genera un modello Excel professionale e formattato, contenente il Bridge contabile
    dall'Utile Netto all'EBITDA Normalizzato utilizzando formule native.
    """
    print(f"\n[Exporter] Generazione del file Excel: {filename}...")

    # Creazione del workbook e selezione del foglio attivo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EBITDA Bridge"

    # Attivazione delle griglie (best practice nei modelli finanziari)
    ws.views.sheetView[0].showGridLines = True

    # --- STILI CORPORATE ---
    font_title = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=11, bold=True)
    font_regular = Font(name="Arial", size=11)
    font_italic = Font(name="Arial", size=9, italic=True, color="595959")

    # Palette Sobria (Navy/Gray) adatta a boutique di advisory
    fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    border_thin = Side(border_style="thin", color="D9D9D9")
    border_double = Side(border_style="double", color="000000")
    border_top_thin = Side(border_style="thin", color="000000")

    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_top_thin, bottom=border_double)

    # --- 1. TITOLO E METADATI ---
    ws.merge_cells("A1:D1")
    ws["A1"] = f"FINANCIAL BRIDGE & EBITDA NORMALIZATION — {consolidated_data['ticker']}"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40

    ws["A2"] = f"Esercizio Fiscale: {consolidated_data['year']} | Data Chiusura: {consolidated_data['end_date']}"
    ws["A2"].font = font_italic
    ws.row_dimensions[2].height = 20

    # --- 2. INTESTAZIONI DI TABELLA ---
    headers = ["Voce di Bilancio / Aggiustamento", "Valore (USD mln)", "Fonte Dati",
               "Note di Riconciliazione / Giustificazione"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 1 else align_left
        cell.border = box_border
    ws.row_dimensions[4].height = 26

    # --- 3. SEZIONE QUANTITATIVA (SEC XBRL) ---
    q_base = consolidated_data["quantitative_base"]

    # Mappiamo i valori convertendoli in milioni per renderli leggibili
    rows_quantitative = [
        ("Net Income (Loss)", q_base.get("Net Income", 0) / 1_000_000, "SEC XBRL API"),
        ("Provision for Income Taxes", q_base.get("Taxes", 0) / 1_000_000, "SEC XBRL API"),
        ("Interest Expense", q_base.get("Interest Expense", 0) / 1_000_000, "SEC XBRL API"),
        ("Depreciation & Amortization (D&A)", q_base.get("D&A", 0) / 1_000_000, "SEC XBRL API")
    ]

    current_row = 5
    for desc, val, fonte in rows_quantitative:
        ws.cell(row=current_row, column=1, value=desc).font = font_regular
        ws.cell(row=current_row, column=2, value=val).font = font_regular
        ws.cell(row=current_row, column=2).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=3, value=fonte).font = font_regular
        ws.cell(row=current_row, column=3).alignment = align_center

        # Applica bordi leggeri
        for c in range(1, 5):
            ws.cell(row=current_row, column=c).border = Border(bottom=border_thin)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # --- 4. RIGA UNADJUSTED EBITDA (FORMULA NATIVA) ---
    ws.cell(row=current_row, column=1, value="UNADJUSTED EBITDA (Base)").font = font_bold
    # Inseriamo la formula Excel nativa per la somma delle righe precedenti (da riga 5 a riga 8)
    formula_unadjusted = f"=SUM(B5:B{current_row - 1})"
    cell_unadj_val = ws.cell(row=current_row, column=2, value=formula_unadjusted)
    cell_unadj_val.font = font_bold
    cell_unadj_val.number_format = "$#,##0.00"

    ws.cell(row=current_row, column=3, value="Calcolo Matematico").font = font_bold
    ws.cell(row=current_row, column=3).alignment = align_center

    for c in range(1, 5):
        cell = ws.cell(row=current_row, column=c)
        cell.fill = fill_total
        cell.border = total_border
    ws.row_dimensions[current_row].height = 24

    row_unadjusted_ebitda = current_row
    current_row += 2  # Lasciamo una riga vuota di spaziatura strutturale

    # --- 5. SEZIONE QUALITATIVA (AGGIUSTAMENTI AI LAYER) ---
    ws.cell(row=current_row, column=1, value="Aggiustamenti Strutturali / Voci Straordinarie:").font = font_bold
    current_row += 1

    start_adj_row = current_row
    adjustments = consolidated_data.get("adjustments", [])

    if not adjustments:
        # Se non ci sono aggiustamenti estratti, inseriamo una riga neutra per non rompere le formule
        ws.cell(row=current_row, column=1, value="Nessun aggiustamento non ricorrente rilevato").font = font_regular
        ws.cell(row=current_row, column=2, value=0.00).font = font_regular
        ws.cell(row=current_row, column=2).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=3, value="-").font = font_regular
        ws.cell(row=current_row, column=3).alignment = align_center
        current_row += 1
    else:
        for adj in adjustments:
            ws.cell(row=current_row, column=1,
                    value=adj.get("adjustment_type", "Voce Straordinaria")).font = font_regular

            # Converte l'importo (assicurandoci che sia un float)
            amt = float(adj.get("amount_millions", 0))
            ws.cell(row=current_row, column=2, value=amt).font = font_regular
            ws.cell(row=current_row, column=2).number_format = "$#,##0.00"

            ws.cell(row=current_row, column=3, value="Claude AI Layer (Item 7)").font = font_regular
            ws.cell(row=current_row, column=3).alignment = align_center

            ws.cell(row=current_row, column=4, value=adj.get("justification", "")).font = font_regular

            for c in range(1, 5):
                ws.cell(row=current_row, column=c).border = Border(bottom=border_thin)
            ws.row_dimensions[current_row].height = 22
            current_row += 1

    end_adj_row = current_row - 1
    current_row += 1

    # --- 6. RIGA NORMALIZED EBITDA (FORMULA NATIVA COMBINATA) ---
    ws.cell(row=current_row, column=1, value="NORMALIZED EBITDA").font = font_bold
    # Formula nativa: Unadjusted EBITDA + Somma degli Aggiustamenti AI
    formula_normalized = f"=B{row_unadjusted_ebitda}+SUM(B{start_adj_row}:B{end_adj_row})"
    cell_norm_val = ws.cell(row=current_row, column=2, value=formula_normalized)
    cell_norm_val.font = font_bold
    cell_norm_val.number_format = "$#,##0.00"

    ws.cell(row=current_row, column=3, value="EBITDA Normalizzato").font = font_bold
    ws.cell(row=current_row, column=3).alignment = align_center

    for c in range(1, 5):
        cell = ws.cell(row=current_row, column=c)
        cell.fill = fill_title
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.border = total_border
    ws.row_dimensions[current_row].height = 26

    # --- 7. AUTO-FIT DELLE COLONNE ---
    # Ottimizzazione visiva per evitare i fastidiosi "###" di Excel su colonne strette
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:  # Ignora il titolo gigante unito per il calcolo della larghezza
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        # Imposta una larghezza minima adattiva con un margine di sicurezza
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Forza una larghezza specifica per la colonna delle giustificazioni per renderla leggibile
    ws.column_dimensions["D"].width = 50

    # Salvataggio effettivo del file
    wb.save(filename)
    print(f"[Exporter] Report salvato con successo.")